import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime


class TelegramBot:
    # создание бота
    def __init__(self, token):

        self.token = token
        self.url = f"https://api.telegram.org/bot{token}/"

    # получение обновлений (все последние сообщения)
    def get_updates(self):

        response = requests.get(f"{self.url}getUpdates")
        data = response.json()
        return data

    # получить текст последнего сообщения
    def get_last(self):

        data = self.get_updates()
        return data["result"][-1]["message"]["text"]

    # отправка сообщения
    def send_message(self, chat_id, text):

        requests.get(f"{self.url}sendMessage?chat_id={chat_id}&text={text}")

    # адресная справка
    def sync_chats(self):

        data = self.get_updates()
        chats = {}

        for update in data["result"]:
            key = update["message"]["chat"]["id"]
            value = update["message"]["from"]["first_name"]
            chats[key] = value
        self.chats = chats
        return chats

    def spam(self, text):

        for chat, name in self.chats.items():
            text_message = f"Hello, {name}! {text}"
            self.send_message(chat, text_message)



    # создания excel с чатами
    def create_chats_excel(self, file_name=None):

        if not file_name:
            # генерируем название
            now = int(datetime.now().timestamp())
            file_name = f"chats{now}.xlsx"

        self.file_name = file_name
        # создаём новый excel файл
        new_excel = Workbook()
        page = new_excel.active
        page["A1"] = "ID"
        page["B1"] = "Имя"

        # записываем чаты в excel
        chats = self.sync_chats()
        for key, value in chats.items():
            page.append([key, value])

        # сохраняем
        new_excel.save(file_name)
        return file_name

    # чтение записей с excel
    def get_chats_excel(self):

        excel_file = load_workbook(self.file_name)
        page = excel_file[excel_file.sheetnames[0]]
        e_read = len(page["A"])
        chats = {}

        for row in range(1, e_read+1):
            chat_id = page[f"A{row}"].value
            name = page[f"B{row}"].value
            chats[chat_id] = name
        return chats    
    
    #чтение с excel spam
    def spam_f_excel(self, text_e):
        
        excel_file = load_workbook(self.file_name)
        page_1 = excel_file[excel_file.sheetnames[0]]
        end_row = len(page_1["A"])

        for row_0 in range(2, end_row + 1):
            mark_1 = page_1[f"A{row_0}"].value
            mark_2 = page_1[f"B{row_0}"].value
            text_message_excel = f"Hello, {mark_2}! {text_e}"
            chat_id_0 = mark_1
            self.send_message(chat_id_0,text_message_excel)
        

    #отправка фото по url
    def send_photo(self, chat_id, photo_url):
        requests.get(f"{self.url}sendPhoto?chat_id={chat_id}&photo={photo_url}")

    #отправка стрикера по id
    def get_sticker(self, chat_id, sticker_id):
        requests.get(f"{self.url}sendSticker?chat_id={chat_id}&sticker={sticker_id}")