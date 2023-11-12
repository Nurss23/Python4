from openpyxl import load_workbook

file_name = "semester.xlsx"
excel = load_workbook(file_name)
list_1 = excel["Лист1"]

list_1["A1"] = "Имя"
list_1["B1"] = "Возраст"
list_1["C1"] = "Факультет"
list_1["D1"] = "Специальность"
list_1["E1"] = "Оценки за семестр"
list_1["F1"] = "Средняя оценка"

data = [
    ["Айгуль Исабекова", 20, "Инженерия", "Информатика",[90, 85, 78, 92],0],
    ["Эркин Торогай", 21, "Медицина", "Медсестринство",[88, 92, 75, 86],0],
    ["Чолпон Сагынбаева", 22, "Инженерия" , "Механика",[86, 90, 87, 94],0],
    ["Бегимай Таалай", 19, "Искусства", "Живопись",[75, 82, 89, 91], 0],
    ["Азамат Абдымомунова", 20, "Биология", "Экология",[91, 88, 84, 90], 0]
]

for i, e in enumerate(data, start = 1):
    for j,k in enumerate(e):
        e[-1] = (sum(e[-2]) / len(e[-2]))
        e[-1] = round(e[-1], 2)
        if isinstance(k,list):
            k = ', '.join(map(str,k))
        list_1.cell(row=i+1,column=j+1,value=k)
excel.save(file_name)


# for i, e in enumerate(data,start= 2):
#     for j,k in enumerate(e,):
        
#         grades = e[-2]
#         list_1[f"A{i}"] = k
#         list_1[f"B{i}"] = k
#         list_1[f"C{i}"] = k
#         list_1[f"D{i}"] = k
#         list_1[f"E{i}"] = ', '.join(map(str,grades))
#         list_1[f"F{i}"] = k

