import requests
from openpyxl import load_workbook

file_name = "posts5.xlsx"
excel = load_workbook(file_name)
list_1 = excel["Лист1"]
list_1["A1"] = "title"
list_1["B1"] = "body"

postss = requests.get("https://jsonplaceholder.typicode.com/posts")
posts = postss.json()

for i,e in enumerate(posts,start =2):
    if len(list_1["A"]) == 11:
        break
    list_1[f"A{i}"] = e["title"]
    list_1[f"B{i}"] = e["body"]
for i, e in enumerate(posts,start=2):
    if len(list_1["B"]) == 11:
        break
    list_1.cell(row=i,column=1,value=e["title"])
    list_1.cell(row=i,column=2,value=e["body"])
excel.save(file_name)